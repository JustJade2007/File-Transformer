"""Data and tabular converter for CSV, TSV, JSON, XML, YAML, XLSX, and SQLite."""
import csv
import json
import os
import sqlite3
import time
from typing import Any, Callable, Dict, List, Optional, Set
import xml.etree.ElementTree as ET
import yaml
import openpyxl

from .base import BaseConverter, ConversionResult

DATA_FORMATS = {"csv", "tsv", "json", "xml", "yaml", "yml", "xlsx", "sqlite", "db", "sql"}


class DataConverter(BaseConverter):
    """Handles structured data, tabular formats, spreadsheets, and database conversions."""

    name = "Data & Serialization Engine"
    category = "Data"

    def supported_inputs(self) -> Set[str]:
        return DATA_FORMATS

    def supported_outputs(self, input_ext: Optional[str] = None) -> Set[str]:
        if not input_ext:
            return {"csv", "tsv", "json", "xml", "yaml", "xlsx", "sql", "sqlite"}
        clean = input_ext.lower().lstrip(".")
        all_targets = {"csv", "tsv", "json", "xml", "yaml", "xlsx", "sql", "sqlite"}
        # Return all other data formats except source itself
        return {t for t in all_targets if t != clean}

    def get_default_options(self, source_ext: str, target_ext: str) -> Dict:
        return {
            "indent": 2,
            "include_headers": True,
            "table_name": "data_table",
        }

    def _read_data_to_records(self, source_path: str, source_ext: str) -> List[Dict[str, Any]]:
        """Parse source file into a list of row dictionaries."""
        records = []

        if source_ext in ("csv", "tsv"):
            delimiter = "\t" if source_ext == "tsv" else ","
            with open(source_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                for row in reader:
                    records.append(dict(row))

        elif source_ext == "json":
            with open(source_path, "r", encoding="utf-8") as f:
                raw = json.load(f)
                if isinstance(raw, list):
                    records = [r if isinstance(r, dict) else {"value": r} for r in raw]
                elif isinstance(raw, dict):
                    records = [raw]

        elif source_ext in ("yaml", "yml"):
            with open(source_path, "r", encoding="utf-8") as f:
                raw = yaml.safe_load(f)
                if isinstance(raw, list):
                    records = [r if isinstance(r, dict) else {"value": r} for r in raw]
                elif isinstance(raw, dict):
                    records = [raw]

        elif source_ext == "xlsx":
            wb = openpyxl.load_workbook(source_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
                for r in rows[1:]:
                    row_dict = {}
                    for i, val in enumerate(r):
                        key = headers[i] if i < len(headers) else f"col_{i}"
                        row_dict[key] = val
                    records.append(row_dict)

        elif source_ext in ("sqlite", "db"):
            conn = sqlite3.connect(source_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
            tables = cursor.fetchall()
            if tables:
                table_name = tables[0][0]
                cursor.execute(f'SELECT * FROM "{table_name}"')
                col_names = [desc[0] for desc in cursor.description]
                for row in cursor.fetchall():
                    records.append(dict(zip(col_names, row)))
            conn.close()

        elif source_ext == "xml":
            tree = ET.parse(source_path)
            root = tree.getroot()
            for child in root:
                row_dict = {}
                for sub in child:
                    row_dict[sub.tag] = sub.text
                if not row_dict and child.text:
                    row_dict[child.tag] = child.text
                records.append(row_dict)

        return records

    def _write_records(
        self,
        records: List[Dict[str, Any]],
        target_path: str,
        target_ext: str,
        options: Dict
    ):
        """Write list of records to the target format."""
        if not records:
            # Handle empty records safely
            records = [{"status": "empty"}]

        # Gather uniform fieldnames
        headers = []
        for r in records:
            for k in r.keys():
                if k not in headers:
                    headers.append(k)

        if target_ext in ("csv", "tsv"):
            delimiter = "\t" if target_ext == "tsv" else ","
            with open(target_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=headers, delimiter=delimiter)
                writer.writeheader()
                for r in records:
                    writer.writerow(r)

        elif target_ext == "json":
            indent = int(options.get("indent", 2))
            with open(target_path, "w", encoding="utf-8") as f:
                json.dump(records, f, indent=indent, default=str)

        elif target_ext in ("yaml", "yml"):
            with open(target_path, "w", encoding="utf-8") as f:
                yaml.dump(records, f, default_flow_style=False, sort_keys=False)

        elif target_ext == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(headers)
            for r in records:
                ws.append([r.get(h, "") for h in headers])
            wb.save(target_path)

        elif target_ext == "xml":
            root = ET.Element("dataset")
            for r in records:
                item = ET.SubElement(root, "record")
                for k, v in r.items():
                    safe_key = "".join(c if c.isalnum() else "_" for c in str(k))
                    sub = ET.SubElement(item, safe_key)
                    sub.text = str(v) if v is not None else ""
            tree = ET.ElementTree(root)
            tree.write(target_path, encoding="utf-8", xml_declaration=True)

        elif target_ext in ("sqlite", "db"):
            if os.path.exists(target_path):
                os.remove(target_path)
            conn = sqlite3.connect(target_path)
            cur = conn.cursor()
            table_name = options.get("table_name", "data_table")
            safe_cols = ['"{}" TEXT'.format(h.replace('"', '""')) for h in headers]
            cur.execute(f'CREATE TABLE "{table_name}" ({", ".join(safe_cols)})')
            placeholders = ", ".join(["?"] * len(headers))
            for r in records:
                vals = [str(r.get(h, "")) if r.get(h) is not None else "" for h in headers]
                cur.execute(f'INSERT INTO "{table_name}" VALUES ({placeholders})', vals)
            conn.commit()
            conn.close()

        elif target_ext == "sql":
            table_name = options.get("table_name", "data_table")
            lines = [f"-- Exported by File-Transformer", f"-- Table: {table_name}"]
            safe_cols = ['"{}" TEXT'.format(h.replace('"', '""')) for h in headers]
            lines.append(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(safe_cols)});')
            for r in records:
                escaped_vals = []
                for h in headers:
                    v = r.get(h)
                    if v is None:
                        escaped_vals.append("NULL")
                    else:
                        val_str = str(v).replace("'", "''")
                        escaped_vals.append(f"'{val_str}'")
                cols_str = ', '.join([f'"{h}"' for h in headers])
                lines.append(f'INSERT INTO "{table_name}" ({cols_str}) VALUES ({", ".join(escaped_vals)});')
            with open(target_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines) + "\n")

    def convert(
        self,
        source_path: str,
        target_path: str,
        target_format: str,
        options: Optional[Dict] = None,
        progress_callback: Optional[Callable[[float, str], None]] = None,
        cancel_event: Optional[object] = None,
    ) -> ConversionResult:
        start_time = time.time()
        source_ext = os.path.splitext(source_path)[1].lower().lstrip(".")
        target_ext = target_format.lower().lstrip(".")
        opts = options or self.get_default_options(source_ext, target_ext)

        os.makedirs(os.path.dirname(os.path.abspath(target_path)), exist_ok=True)

        try:
            if progress_callback:
                progress_callback(0.25, f"Parsing {source_ext.upper()} data...")

            records = self._read_data_to_records(source_path, source_ext)

            if cancel_event and getattr(cancel_event, "is_set", lambda: False)():
                return ConversionResult(success=False, error_message="Cancelled by user.")

            if progress_callback:
                progress_callback(0.70, f"Writing {target_ext.upper()} records ({len(records)} rows)...")

            self._write_records(records, target_path, target_ext, opts)

            if progress_callback:
                progress_callback(1.0, "Complete")

            return ConversionResult(
                success=True,
                output_path=target_path,
                duration_seconds=time.time() - start_time,
            )
        except Exception as e:
            return ConversionResult(
                success=False,
                error_message=f"Data conversion failed: {str(e)}",
                duration_seconds=time.time() - start_time,
            )
